import openpyxl
from openpyxl.utils import get_column_letter
import re
from SubjestClass import *
from DataBase import Schedule
from utils import translate_weekday
import UserClass
from threading import *
from time import time, sleep
import os

GroupList = []

subjects = {}
multisubjects = {}

exams = []


def Parser_Table_For_Exams(name_table):
    # name_table = "schedules\КБиСП 2 курс 2 сем-Д (3).xlsx"  # потом может удалить

    start = time()

    group_list_local = []
    table = openpyxl.open(name_table, read_only=True)

    sheet = table.active

    wb = openpyxl.load_workbook(name_table)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]]
    row_max = 3
    column_max = sheet_active.max_column
    # print(column_max, ' ', row_max) # проверка количества строк и столбцов
    row_min = 3
    column_min = 1
    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max

        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        # print(data_from_cell)
        if len(data_from_cell) > 9:
            if data_from_cell[4] == '-' and data_from_cell[7] == '-' and data_from_cell not in group_list_local:
                group_list_local.append(data_from_cell)
                # print(data_from_cell, column_min, " ", row_min_min)
                cells = sheet[get_column_letter(column_min) + '4':get_column_letter(column_min + 3) + '59']
                subjets = []
                day_week = 0
                if cells != None:
                    timez = 0
                    type_par = 0
                    nomer = 0
                    FIO = 0
                    item = 0
                    k = 0
                    day = 0
                    for cell, cell1, cell2, cell3 in cells:
                        day_week += 1
                        if day_week == 1:
                            type_par = cell
                            timez = cell1
                            nomer = cell2
                            if cell.value is not None:
                                day = sheet[cell.row][2]

                        if day_week == 2:
                            item = cell
                        if day_week == 3:
                            FIO = cell

                        if day_week == 3:
                            if type_par.value is not None:
                                #    print(data_from_cell[:10], item.value.strip().replace('\n', ' '),timez.value,FIO.value, nomer.value,2 if 'экз' in name_table.lower() else 0, type_par.value, day.value)
                                exams.append([data_from_cell[:10], day.value.strip().replace('\n', ' '), timez.value,
                                              type_par.value, item.value, FIO.value, nomer.value])
                            k += 1
                            day_week = 0
                            if k == 6:
                                k = 0
                                day_week = -1

        row_min_min = int(row_min_min)

        column_min = column_min + 1
    print(time() - start, name_table)
    wb.close()
    table.close()

def Parser_Table(name_table):

    start = time()

    group_list_local = []
    table = openpyxl.open(name_table, read_only=True)

    sheet = table.active

    wb = openpyxl.load_workbook(name_table)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]]
    row_max = 2
    column_max = sheet_active.max_column
    row_min = 2
    column_min = 1
    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max

        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)

        if len(data_from_cell) > 9:
            if data_from_cell[4] == '-' and data_from_cell[7] == '-' and data_from_cell not in group_list_local:

                group_list_local.append(data_from_cell)

                cells = sheet[get_column_letter(column_min) + '4':get_column_letter(column_min + 4) + '75']
                chetnost = 0
                NomerPar = 2
                DenNedeli = 4
                subjets = []
                for item, vid_zanyatiy, FIO, nomer, ssilka in cells:

                    if NomerPar == 14:
                        NomerPar -= 12

                        subjets.clear()
                        DenNedeli += 12

                    if item.value != None and "\n" in item.value and item.value.strip() != 'Военная\nподготовка':
                        try:
                            ForMultiSubject = MultiSubject(data_from_cell[:10], NomerPar // 2, chetnost % 2,
                                                           item.value.strip(),
                                                           vid_zanyatiy.value,
                                                           FIO.value, nomer.value, ssilka.value,
                                                           1 if 'зач' in name_table.lower() else 0)
                        except Exception as e:
                            print(str(e))
                            a = 1 / (1 - 1)
                        global multisubjects
                        multisubjects[ForMultiSubject] = translate_weekday(sheet[DenNedeli][0].value)


                    elif item.value != None:
                        ForSubject = Subject(data_from_cell[:10], NomerPar // 2, chetnost % 2,
                                             item.value.strip().replace('\n', ' '),
                                             vid_zanyatiy.value,
                                             FIO.value, nomer.value, ssilka.value,
                                             1 if 'зач' in name_table.lower() else 0)
                        global subjects
                        subjects[ForSubject] = translate_weekday(sheet[DenNedeli][0].value)



                    chetnost += 1
                    NomerPar += 1

        row_min_min = int(row_min_min)

        column_min = column_min + 1
    print(time() - start, name_table)
    wb.close()
    table.close()



def Groups_List(name_table):

    table = openpyxl.open(name_table, read_only=True)

    sheet = table.active
    wb = openpyxl.load_workbook(name_table)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]]
    row_max = 2
    column_max = sheet_active.max_column
    # print(column_max, ' ', row_max) # проверка количества строк и столбцов
    row_min = 2
    column_min = 1
    global GroupList

    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max

        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)

        if len(data_from_cell) > 9:
            if data_from_cell[4] == '-' and data_from_cell[7] == '-' and data_from_cell[:10] not in GroupList:


                GroupList.append(data_from_cell[0:10])
        row_min_min = int(row_min_min)

        column_min = column_min + 1
    wb.close()
    table.close()



def SchedulePars(Exams = 0):
    schedule = Schedule(path='private/rasp1.db')
    schedule.drop_tables()
    os.chdir('./schedules')  # переход на работу с другой директорией
    dirlist = ['./schedules/' + el for el in os.listdir()]
    os.chdir('../')
    global GroupList
    GroupList = []

    start = time()
    for NameTable12 in dirlist:  # цикл для работы с списком ссего
        Groups_List(NameTable12)
    GroupList.sort()
    schedule.insert_groups(GroupList)

    if Exams != 0:
        global exams
        exams.clear()
        for NameTable1 in dirlist:  # цикл для работы с списком экзов
            if "экз" or "ekz" in NameTable1:
                Parser_Table_For_Exams(NameTable1)
        schedule.insert_exams(exams)

    else:
        global subjects, multisubjects
        subjects.clear()
        multisubjects.clear()

        for NameTable1 in dirlist:  # цикл для работы с списком всего
            Parser_Table(NameTable1)


        for key, value in multisubjects.items():
            schedule.insert_subjects(value, key.subjects)
        subjects_grouped_by_weekday = {key: [] for key in 'Monday Tuesday Wednesday Thursday Friday Saturday'.split()}
        for key, value in subjects.items():
            subjects_grouped_by_weekday[value].append(key)
        for key, value in subjects_grouped_by_weekday.items():
            schedule.insert_subjects(key, value)
    print(time() - start)
    schedule.con.close()

# SchedulePars()
# schedule.insert_groups(GroupsList())

# SchedulePars(1)